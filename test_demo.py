from datetime import date
from pathlib import Path
import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
from Constants import global_constants


def get_data():
    excel_file = openpyxl.load_workbook("Data/invalid_login.xlsx")
    selected_sheet = excel_file["Sayfa1"]
    total_row = selected_sheet.max_row
    data = list()
    for i in range(2, total_row + 1):
        username = selected_sheet.cell(i, 1).value
        password = selected_sheet.cell(i, 2).value
        tuple_data = (username, password)
        data.append(tuple_data)
    return data


class Test_demo_class:
    # calls before test
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(global_constants.URL)
        self.folder_path = str(date.today())
        Path(self.folder_path).mkdir(exist_ok=True)

    # calls after test
    def teardown_method(self):
        self.driver.quit()

    def test_demo_function(self):
        text = "deneme"
        assert text == "deneme"

    def test_demo2(self):
        assert True

    @pytest.mark.parametrize("username,password", get_data())
    def test_invalid_login(self, username, password):
        self.wait_for_element_visible((By.ID, "user-name"))
        user_name_input = self.driver.find_element(By.ID, "user-name")

        self.wait_for_element_visible((By.ID, "password"), 10)
        password_input = self.driver.find_element(By.ID, "password")

        user_name_input.send_keys(username)
        password_input.send_keys(password)
        login_btn = self.driver.find_element(By.ID, "login-button")
        login_btn.click()
        error_message = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        screenshot_file_path = str(Path(self.folder_path) / f"test-invalid-login-{username}-{password}.png")
        self.driver.save_screenshot(screenshot_file_path)
        assert error_message.text == "Epic sadface: Username and password do not match any user in this service"

    def wait_for_element_visible(self, locator, timeout=5):
        WebDriverWait(self.driver, timeout).until(ec.visibility_of_element_located(locator))
